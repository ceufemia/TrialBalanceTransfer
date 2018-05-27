import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
#Test line
# Source = openpyxl.load_workbook('F:\HITB0916.xlsx')
Target = openpyxl.load_workbook('C:\Users\T530\Desktop\Banta\Banta Management Chart of Account Mapping.xlsx')
TargetSheet = Target.get_sheet_by_name(name='TB from Prior Mgmt')
# SourceSheet = Source.get_sheet_by_name(name='Sheet1')

initials = "GF"
length = 255

Source = openpyxl.load_workbook('C:\Users\T530\Desktop\Banta\\' + initials + '\\' + initials + 'TB201312.xlsx')
SourceSheet = Source.get_sheet_by_name(name='Sheet1')
for i in range(14,length,1):
    j = i-9
    print i
    AccountNumColumn = get_column_letter(1)
    AccountNameColumn = get_column_letter(2)
    DebitColumn = get_column_letter(4)
    CreditColumn = get_column_letter(5)
    NetChangeColumn = get_column_letter(6)
    EndingColumn = get_column_letter(7)

    TargetSheet[AccountNumColumn + str(j)].value = SourceSheet['A' + str(i)].value
    TargetSheet[AccountNameColumn + str(j)].value = SourceSheet['E' + str(i)].value
    TargetSheet[DebitColumn + str(j)].value = SourceSheet['H' + str(i)].value
    TargetSheet[CreditColumn + str(j)].value = SourceSheet['I' + str(i)].value
    if TargetSheet[DebitColumn + str(j)].value or TargetSheet[DebitColumn + str(j)].value == 0:
        TargetSheet[EndingColumn + str(j)].value = TargetSheet[DebitColumn + str(j)].value
    elif TargetSheet[CreditColumn + str(j)].value or TargetSheet[CreditColumn + str(j)].value == 0:
        TargetSheet[EndingColumn + str(j)].value = TargetSheet[CreditColumn + str(j)].value * -1

for index in range(0, column_index_from_string('OR') / 8, 1):

    AccountNumColumn = get_column_letter(index * 8 + 9)
    AccountNameColumn = get_column_letter(index * 8 + 10)
    OpeningColumn = get_column_letter(index * 8 + 11)
    DebitColumn = get_column_letter(index * 8 + 12)
    CreditColumn = get_column_letter(index * 8 + 13)
    NetChangeColumn = get_column_letter(index * 8 + 14)
    EndingColumn = get_column_letter(index * 8 + 15)

    year = 2014+ (index/12)
    month = (index%12)+1
    print str(month) + str(year)


    '''if  month == 0:
        try:
            Source = openpyxl.load_workbook('C:\Users\T530\Desktop\Banta\\' + initials + '\\' + initials + 'TB' + str(year) + '12.xlsx')
        except:
            print "ERROR: " + str(month)+str(year)'''

    if month < 10:
        Source = openpyxl.load_workbook('C:\Users\T530\Desktop\Banta\\'+initials+'\\'+initials+'TB' + str(year) + '0' + str(month) + '.xlsx')

    else:
        Source = openpyxl.load_workbook('C:\Users\T530\Desktop\Banta\\'+initials+'\\'+initials+'TB' +str(year) + str(month) + '.xlsx')

    SourceSheet = Source.get_sheet_by_name(name='Sheet1')

    for i in range(14, length, 1):
        j = i - 9
        TargetSheet[AccountNumColumn + str(j)].value = SourceSheet['A' + str(i)].value
        TargetSheet[AccountNameColumn + str(j)].value = SourceSheet['E' + str(i)].value

        if int(TargetSheet[AccountNumColumn + str(j)].value)  == 1996:
            RetainedRow = j

        TargetSheet[OpeningColumn + str(j)].value = TargetSheet[get_column_letter(index * 8 + 7) + str(j)].value
            #print TargetSheet[OpeningColumn + str(j)].value
            #print TargetSheet[AccountNumColumn + str(j)].value

        if index % 12 == 0 and int(TargetSheet[AccountNumColumn + str(j)].value) > 1999:
            print OpeningColumn
            print RetainedRow
            TargetSheet[OpeningColumn + str(RetainedRow)].value += TargetSheet[get_column_letter(index * 8 + 7) + str(j)].value
            TargetSheet[OpeningColumn + str(j)].value = 0
            #print TargetSheet[OpeningColumn + str(RetainedRow)].value



        TargetSheet[DebitColumn + str(j)].value = SourceSheet['H' + str(i)].value

        TargetSheet[CreditColumn + str(j)].value = SourceSheet['I' + str(i)].value

        if index < 1:
            if TargetSheet['D' + str(i)].value:
                TargetSheet[EndingColumn + str(j)].value = TargetSheet['D' + str(i)].value
            elif TargetSheet['E' + str(i)].value:
                TargetSheet[EndingColumn + str(j)].value = TargetSheet['E' + str(i)].value * -1

        if TargetSheet[DebitColumn + str(j)].value or TargetSheet[DebitColumn + str(j)].value == 0:
            TargetSheet[EndingColumn + str(j)].value = TargetSheet[DebitColumn + str(j)].value
        elif TargetSheet[CreditColumn + str(j)].value or TargetSheet[CreditColumn + str(j)].value == 0:
            TargetSheet[EndingColumn + str(j)].value = TargetSheet[CreditColumn + str(j)].value * -1
        else:
            #print CreditColumn + str(j) + DebitColumn + str(j)
            #print str(TargetSheet[CreditColumn + str(j)].value) + ' ' + str(TargetSheet[DebitColumn + str(j)].value)
            break


            #print NetChangeColumn + str(j) + EndingColumn + str(j) + OpeningColumn + str(j)
        TargetSheet[NetChangeColumn + str(j)].value = TargetSheet[EndingColumn + str(j)].value - TargetSheet[OpeningColumn + str(j)].value

            # Target.save('F:\Banta Management Chart of Account Mapping - MF (3).xlsx')
    #if index > 48:
     #   Target.save('C:\Users\T530\Desktop\Banta\Banta Management Chart of Account Mapping - '+initials+' (1).xlsx')
Target.save('C:\Users\T530\Desktop\Banta\Banta Management Chart of Account Mapping - '+initials+'.xlsx')

Full = openpyxl.load_workbook('C:\Users\T530\Desktop\Banta\Banta Management Chart of Account Mapping.xlsx')
FullSheet = Target.get_sheet_by_name(name='TB from Prior Mgmt')

def SplitYears(Fullfile, Yearfile, start, end, initials, year):
     Yearbook = openpyxl.load_workbook(Yearfile)
     Fullbook = openpyxl.load_workbook(Fullfile)
     Yearsheet = Yearbook.get_sheet_by_name(name='TB from Prior Mgmt')
     Fullsheet = Fullbook.get_sheet_by_name(name='TB from Prior Mgmt')
     Fullrange = range(column_index_from_string(start),column_index_from_string(end)+3)


     for index in range(1, 105):
         for j in range (4, 279):
             Full = Fullrange[index-1]

             Yearsheet[get_column_letter(index)+str(j-2)].value = Fullsheet[get_column_letter(Full)+str(j)].value

     Yearbook.save('C:\Users\T530\Desktop\Banta\\'+initials+'\\'+initials+year+' Banta Management Chart of Account Mapping .xlsx')


SplitYears('C:\Users\T530\Desktop\Banta\Banta Management Chart of Account Mapping - '+initials+'.xlsx',
     'C:\Users\T530\Desktop\Banta\Banta Management Chart of Account Mapping 2014.xlsx', 'A', 'CZ', initials, '2014')
SplitYears('C:\Users\T530\Desktop\Banta\Banta Management Chart of Account Mapping - '+initials+'.xlsx',
     'C:\Users\T530\Desktop\Banta\Banta Management Chart of Account Mapping 2015.xlsx', 'CR', 'GS', initials, '2015')
SplitYears('C:\Users\T530\Desktop\Banta\Banta Management Chart of Account Mapping - '+initials+'.xlsx',
     'C:\Users\T530\Desktop\Banta\Banta Management Chart of Account Mapping 2016.xlsx', 'GK', 'KK', initials, '2016')
SplitYears('C:\Users\T530\Desktop\Banta\Banta Management Chart of Account Mapping - '+initials+'.xlsx',
     'C:\Users\T530\Desktop\Banta\Banta Management Chart of Account Mapping 2017.xlsx', 'KC', 'OC', initials, '2017')
SplitYears('C:\Users\T530\Desktop\Banta\Banta Management Chart of Account Mapping - '+initials+'.xlsx',
     'C:\Users\T530\Desktop\Banta\Banta Management Chart of Account Mapping 2018.xlsx', 'NU', 'RS', initials, '2018')